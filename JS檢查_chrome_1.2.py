# -*- coding: utf-8 -*-

from selenium import webdriver
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,ytFuntion

#新增帳號讀取

def sheet_date():
    sheet["F" + str(i)].value = time.strftime("%y_%m_%d") #檢查日期
    sheet["G" + str(i)].value = time.strftime("%H_%M_%S") #檢查時間
    
webDriver = webdriver.Chrome(executable_path='chromedriver.exe')
test_web = ytFuntion.test_web(webDriver)

wb = load_workbook("檢查JS用.xlsx")
sheet = wb["web"] # 獲取一張表

testWebUrl = input("請輸入測試站點的url(Ex.http://winvip66.acgtest.com):").strip()
sheet["A1"].value = testWebUrl
webDriver.get(str(testWebUrl) + str(sheet["D2"].value).strip())
testSiteID = input("請輸入測試SiteID(不修改請直接enter):").strip()
sheet["A2"].value = testSiteID
j = 0
for i in range(8,13): #找siteID所在位置
    try:
        int(test_web["site.config"][test_web["site.config"].index("siteId") + i])
        j += 1
    except:
        pass
try:
    newSiteConfig = test_web["site.config"][:test_web["site.config"].index("siteId") + 8] + str(testSiteID) + test_web["site.config"][test_web["site.config"].index("siteId") + 8 + j:]
except:
    input("url錯誤,請檢查並重新啟動。")
    
    
if len(testSiteID.strip()) == 0:
    pass
else:
    test_web["site.config"] = newSiteConfig #更換siteConfig
textCheck = input("請輸入要檢測的字,若開頭為「http」,請輸入到「.com」即可(多個請用空白隔開):").strip()
sheet["A3"].value = textCheck
textCheck = textCheck.split(" ")
textCheckList = []
for i in textCheck:
    if i.strip() != "":
        textCheckList.append(i)
    
print("檢測目標:" + str(textCheckList))
print()

for i in range(2 ,len(sheet["B"])+1):
    if i == 10:
        test_web.elementClick("亲，请登录",3)
        test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = str(sheet["K1" ].value).strip()) #帳密
        test_web.elementSendKeys("input[tag=密码]" ,6 ,text = str(sheet["M1" ].value).strip())
        test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']" ,6)
        sleep(10)
    testUrl = str(testWebUrl) + str(sheet["D" + str(i)].value).strip()
    sheet["D" + str(i)].value = testUrl
    webDriver.get(testUrl)
    webDriver.refresh()
    sleep(3)
    html_source = webDriver.page_source
    if webDriver.current_url == str(testUrl) and "您所访问的彩种不存在，即将返回购彩大厅" not in html_source and "Unexpected token u in JSON at position 0" not in html_source:
        for j in range(len(textCheckList)):
            if textCheckList[j] in html_source:
                sheet.cell(row =i , column = 5 + j ).value = "有"
                #sheet_date()
            else:
                sheet.cell(row =i , column = 5 + j ).value = "沒有"
                #sheet_date()
    else:
        sheet["E" + str(i)].value = "無此url"
        sheet_date()
        
wb.save(str(testSiteID) + "_" + "JS檢查報告_chrome_" + str(time.strftime("%y_%m_%d_%H_%M_%S") + ".xlsx"))
webDriver.quit()
