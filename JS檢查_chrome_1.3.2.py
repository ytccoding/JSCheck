# -*- coding: utf-8 -*-

from selenium import webdriver
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,ytFuntion

#1.3.2,加入等待時間
def sheet_date():
    sheet["F" + str(i)].value = time.strftime("%y_%m_%d") #檢查日期
    sheet["G" + str(i)].value = time.strftime("%H_%M_%S") #檢查時間

groupNumber = input("請輸入Web測試站點序列號:").strip()
   
webDriver = webdriver.Chrome(executable_path='chromedriver.exe')
test_web = ytFuntion.test_web(webDriver)

wbAccount = load_workbook("主副域名對照表.xlsx")
sheetAccount = wbAccount["Account"] # 獲取一張表

wb = load_workbook("檢查JS用.xlsx")
sheet = wb["web"] # 獲取一張表

for i in range(1,len(sheetAccount["A"])+1):
    if str(sheetAccount["A" + str(i)].value).strip() == str(groupNumber):
        testWebUrl = str(sheetAccount["D" + str(i)].value).strip()
        sheet["A1"].value = testWebUrl
        webDriver.get(str(testWebUrl) + str(sheet["D2"].value).strip())
        testSiteID = str(sheetAccount["G" + str(i)].value).strip()
        sheet["A2"].value = testSiteID
        account = str(sheetAccount["E" + str(i)].value).strip()
        sheet["K1"].value = account
        password = str(sheetAccount["F" + str(i)].value).strip()
        sheet["M1"].value = password
        textCheck = str(sheetAccount["H" + str(i)].value).strip()
        if str(sheetAccount["I" + str(i)].value).strip() == "None":
            waitSec = 10
        else:
            waitSec = int(str(sheetAccount["I" + str(i)].value).strip())
sleep(10)    
j = 0
for i in range(8,13): #找siteID所在位置
    try:
        int(test_web["site.config"][test_web["site.config"].index("siteId") + i])
        j += 1
    except:
        pass
try:
    newSiteConfig = test_web["site.config"][:test_web["site.config"].index("siteId") + 8] + str(testSiteID) + test_web["site.config"][test_web["site.config"].index("siteId") + 8 + j:]
except:
    input("url錯誤,請檢查並重新啟動。")
    

if len(testSiteID.strip()) == 0 or testSiteID.strip() == "None":
    pass
else:
    test_web["site.config"] = newSiteConfig #更換siteConfig

sheet["A3"].value = textCheck
textCheck = textCheck.split(" ")
textCheckList = []
for i in textCheck:
    if i.strip() != "":
        textCheckList.append(i)
    
print("檢測目標:" + str(textCheckList))
print()

for i in range(2 ,len(sheet["B"])+1):
    if i == 11:
        testUrl = str(testWebUrl) + str(sheet["D" + str(i)].value).strip()
        webDriver.get(testUrl)
        test_web.elementClick("亲，请登录",3)
        test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = account) #帳密
        test_web.elementSendKeys("input[tag=密码]" ,6 ,text = password)
        test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']" ,6)
        timeCount = 0
        while(test_web.webDriver.current_url != str(testWebUrl) + str(sheet["D" + str(i)].value).strip()):
            sleep(1)
            timeCount = timeCount + 1
            if timeCount >= 30:
                break
    testUrl = str(testWebUrl) + str(sheet["D" + str(i)].value).strip()
    sheet["D" + str(i)].value = testUrl
    webDriver.get(testUrl)
    webDriver.refresh()
    sleep(waitSec)#等待時間
    html_source = webDriver.page_source
    if webDriver.current_url == str(testUrl) and "您所访问的彩种不存在，即将返回购彩大厅" not in html_source and "Unexpected token u in JSON at position 0" not in html_source:
        for j in range(len(textCheckList)):
            if textCheckList[j] in html_source:
                sheet.cell(row =i , column = 5 + j ).value = "有"
                if html_source.count(textCheckList[j]) > 1:
                    input("JS重複出現,按Enter離開。")
                    webDriver.quit()
            else:
                sheet.cell(row =i , column = 5 + j ).value = "沒有"
    else:
        sheet["E" + str(i)].value = "無此url"
        sheet_date()
        
wb.save(str(testSiteID) + "_" + "JS檢查報告_chrome_" + str(time.strftime("%y_%m_%d_%H_%M_%S") + ".xlsx"))
webDriver.quit()

